import os

from django.core.management.base import BaseCommand, CommandError
from core.models import CelestialBody
from openpyxl import load_workbook

class Command(BaseCommand):
    help = 'Populate DataBase with data extracted from excel sheet'

    def add_arguments(self, parser):
        parser.add_argument('excel_sheet', type=str)
    
    def handle(self, *args, **kwargs):

        # Read excel file and opening the active sheet
        file = kwargs['excel_sheet']
        
        wb = load_workbook(os.path.join(os.environ.get('SOLARSYSTEMAPI_ROOT'), file))
        ws = wb.active
        
        # Iterate through the rows and get the values MASS, AGE, RADIUS, BODY_TYPE
        for row in ws.iter_rows(min_row=2, max_row=11,values_only=True):
            (name, mass, radius, age, body_type) = row
            self.stdout.write('name: {}Mass: {}\nRadius: {}\nAge: {}\nBody Type: {}\n\n'.format(name, eval(mass), float(radius), float(age), body_type))
            body = CelestialBody(name=name, mass=eval(mass), radius=float(radius), age=float(age), body_type=body_type)
            body.save()
        

        self.stdout.write("Operation done successfully.", ending='\n')
    
